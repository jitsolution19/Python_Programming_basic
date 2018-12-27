#Write a program to create an Excel file and enter data
import openpyxl
wb = openpyxl.load_workbook("sample.xlsx")
# grab the active worksheet
ws = wb.active
data = ws.cell(row =2,column =1).value
print(data)

# Output file
from openpyxl import Workbook
wb_output = Workbook()
ws_output = wb_output.active
ws_output.cell(row =1,column =1).value = data
wb_output.save("output.xlsx")
